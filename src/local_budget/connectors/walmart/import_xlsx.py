"""A purchase-history spreadsheet export → the entity dicts `store.py` writes.

The second way into this connector, and the one that actually works at scale.
`fetch.py` walks Walmart's own pages, which is fine for the last few orders and
useless for the last few years: PerimeterX challenges list paging after ~2 pages,
so the scraper structurally cannot backfill. A spreadsheet exported from an
already-signed-in browser has no bot-detection surface at all — it is a file.

So the division of labour is: **`sync` keeps the tail fresh, `import` supplies the
history.** Both land in `store.store_orders`, which upserts by order number, so
running one after the other is safe in either order and neither invents a row the
other has to reconcile.

**What this module does NOT read.** The export carries a `Ship To` column holding a
real name and street address. It is deliberately never parsed. Nothing downstream
asks who an order shipped to, and a column no one reads is one that cannot leak
into a report, a fixture, or a commit.

**Money is passed on as the string the sheet displayed**, exactly as `parse.py`
does, so `store.to_cents` keeps using the strict non-rounding conversion path.
The cells are floats, so `repr` would be wrong here — they are formatted back to
two decimals, which is what the spreadsheet showed a human.

**The item rows need judgment, and this is where it is applied.** An export row is
not always a distinct purchase:

- `Canceled` / `Unavailable` lines were never paid for. Dropped.
- `Shopped` and `N weight adjusted` lines are *restatements* — a second listing of
  a line that already appears, mostly on in-store receipts. Dropped ONLY when a
  fulfilled row in the same order carries the same title and price, because the
  same statuses also appear on rows that are the only listing of a real item.
  Blanket-dropping them loses real purchases; blanket-keeping them double-counts
  a grocery line, which surfaces as a visibly wrong category total.

Even after that, item lines do not sum to Walmart's own subtotals — median gap is
about 3%, from checkout-time price adjustments the export does not restate. That
is a property of the source, not of this parser, and it is recorded per order as
`item_sum_cents` rather than smoothed away. It does not corrupt attribution:
`split.propose` scales lines to the actual bank charge before anything is
categorized.
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from ... import db
from . import match, store


class WalmartParseError(RuntimeError):
    """The file did not contain the shape this parser knows.

    Raised rather than returning empty. A silent zero is the failure this
    connector refuses to have: the format changes, the parser yields nothing,
    the import reports success, and every command afterwards prints a confident
    empty table.
    """

#: Sheet names, and the columns each must carry. Checked up front so a renamed
#: header fails loudly on arrival instead of yielding an order with no items —
#: the silent-zero failure this connector refuses to have.
ORDER_SHEET = "Orders"
ITEM_SHEET = "Items"
REQUIRED_ORDER_COLS = ("Order Number", "Order Date", "Order Type", "Order Total")
REQUIRED_ITEM_COLS = ("Order Number", "Product Name", "Qty", "Price", "Status")

#: Item statuses meaning "not purchased". These lines are real history but never
#: money, so they are dropped rather than stored with a price the ledger would
#: have to explain.
DEAD_STATUSES = frozenset({"canceled", "cancelled", "unavailable"})

#: Item statuses that mark a RESTATEMENT of another line rather than a line of
#: its own. Only suppressed when the same (title, price) is already present under
#: a fulfilment status — see the module docstring.
_RESTATEMENT = re.compile(r"^(shopped|\d+\s+weight adjusted|weight adjusted)$", re.I)

#: Walmart's own order-type flag. `GLASS` is its name for the online stack; an
#: in-store receipt says so outright. This feeds `match.patterns_for`, which uses
#: it to keep a grocery pickup from claiming a same-total Supercenter charge.
_CHANNELS = {"GLASS": "online", "IN_STORE": "in-store"}

#: `https://www.walmart.com/ip/<slug>/12345678` or `.../ip/12345678` — the item
#: number is the trailing path segment, and it is the same identifier the scraper
#: stores as `usItemId`, so both sources key products the same way.
_PRODUCT_ID = re.compile(r"/(\d+)/?$")

#: "Jul 01, 2026", and the in-store variant "Apr 16, 2026 purchase".
_DATE_SUFFIX = re.compile(r"\s+purchase$", re.I)


def _text(v) -> str:
    return "" if v is None else str(v).strip()


def parse_date(value) -> str | None:
    """A sheet date cell → ISO date, or None.

    Two formats appear in one column: online orders read "Jul 01, 2026" and
    in-store receipts read "Apr 16, 2026 purchase". The suffix is stripped rather
    than special-cased per sheet, because both kinds share the column.
    """
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = _DATE_SUFFIX.sub("", _text(value))
    if not s:
        return None
    try:
        return datetime.strptime(s, "%b %d, %Y").date().isoformat()
    except ValueError:
        return None


def money_str(value) -> str | None:
    """A numeric cell → the decimal string the sheet displayed.

    Formatted to two places rather than passed through `str()`: the cells are
    floats, so `str(5.6)` would hand "5.6" to a converter documented to read what
    a human read. Blank stays blank — `store.to_cents` maps it to NULL, which is
    the honest value for a field the export omitted.
    """
    if value is None or value == "":
        return None
    if isinstance(value, str):
        return value.strip() or None
    return f"{float(value):.2f}"


def _sum_str(*values) -> str | None:
    """Several money cells → one decimal string, or None if all are blank.

    Walmart bills delivery and the bag fee as separate lines; the schema has one
    `shipping_cents`. Summing here keeps the arithmetic in the parser, where the
    source shape is still visible.
    """
    present = [v for v in values if v not in (None, "")]
    if not present:
        return None
    return f"{sum(float(v) for v in present):.2f}"


def product_id(url: str | None) -> str | None:
    """The Walmart item number out of a product URL."""
    if not url:
        return None
    m = _PRODUCT_ID.search(url.split("?")[0])
    return m.group(1) if m else None


def quantity(value):
    """A qty cell → a number, preserving weights.

    Deli meat is sold by the pound and the export says so: `0.514`. Coercing that
    to an int — which `store._store_items` used to do — reads it as ZERO, which
    silently deletes the line's quantity while keeping its price. Blank means the
    export omitted it, and 1 is the only sane reading of a line that exists.
    """
    if value in (None, ""):
        return 1
    try:
        n = float(value)
    except (TypeError, ValueError):
        return 1
    return int(n) if n == int(n) else n


def _is_dead(status: str) -> bool:
    return status.strip().lower() in DEAD_STATUSES


def select_items(rows: list[dict]) -> list[dict]:
    """One order's raw item rows → the lines that represent real purchases.

    Drops non-purchases outright, then suppresses a restatement only when the
    line it restates is already present. The `used` set makes that one-for-one:
    two genuine copies of the same product are not collapsed by a single
    restatement of it.
    """
    live = [r for r in rows if not _is_dead(r["status"])]
    fulfilled = {(r["title"], r["price"]) for r in live
                 if not _RESTATEMENT.match(r["status"].strip())}
    out: list[dict] = []
    used: set = set()
    for r in live:
        key = (r["title"], r["price"])
        if _RESTATEMENT.match(r["status"].strip()) and key in fulfilled and key not in used:
            used.add(key)
            continue
        out.append(r)
    return out


def _headers(ws, required: tuple[str, ...], sheet: str) -> dict:
    head = next(ws.iter_rows(values_only=True), None)
    if not head:
        raise WalmartParseError(f"sheet '{sheet}' is empty")
    index = {_text(h): i for i, h in enumerate(head) if h is not None}
    missing = [c for c in required if c not in index]
    if missing:
        raise WalmartParseError(
            f"sheet '{sheet}' is missing column(s) {missing} — the export format "
            f"changed, or this is not a Walmart order export")
    return index


def _read_items(ws) -> dict[str, dict]:
    """The Items sheet → ``{order_number: {"items": [...], "all_dead": bool}}``.

    `all_dead` is judged on the RAW rows, before selection. It has to be: an
    order whose every line was cancelled selects down to no items at all, and
    "no items" is indistinguishable from an order this sheet simply did not
    list. The distinction matters — `backfill.pending_detail` skips cancelled
    orders, so getting it wrong sends the scraper after pages forever.
    """
    ix = _headers(ws, REQUIRED_ITEM_COLS, ITEM_SHEET)
    link_col = ix.get("Product Link")
    raw: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2):
        cells = [c.value for c in row]
        number = _text(cells[ix["Order Number"]])
        if not number:
            continue
        link = row[link_col] if link_col is not None and link_col < len(row) else None
        raw.setdefault(number, []).append({
            "title": _text(cells[ix["Product Name"]]) or None,
            "price": cells[ix["Price"]],
            "qty": cells[ix["Qty"]],
            "status": _text(cells[ix["Status"]]),
            "url": link.hyperlink.target if link is not None and link.hyperlink else None,
        })

    return {number: {
        "items": [{
            "title": r["title"],
            "product_id": product_id(r["url"]),
            "quantity": quantity(r["qty"]),
            "line_price": money_str(r["price"]),
            # The export names no seller per line — only an order-level
            # "Seller(s)" list, which cannot be attributed to a specific item
            # without guessing.
            "seller": None,
            # Walmart publishes no shelf taxonomy here either, so the report
            # falls back to the keyword heuristic it is written for.
            "category": None,
            "status": r["status"] or None,
        } for r in select_items(rows)],
        "all_dead": all(_is_dead(r["status"]) for r in rows),
    } for number, rows in raw.items()}


def _read_orders(ws, items: dict[str, dict]) -> list[dict]:
    ix = _headers(ws, REQUIRED_ORDER_COLS, ORDER_SHEET)

    def cell(row, name):
        i = ix.get(name)
        return None if i is None or i >= len(row) else row[i]

    orders = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = _text(cell(row, "Order Number"))
        if not number:
            continue
        found = items.get(number, {})
        # "Before Savings" is the figure the item lines were priced at, so it is
        # the one worth comparing them against. Not every row carries it.
        subtotal = (money_str(cell(row, "Subtotal (Before Savings)"))
                    or money_str(cell(row, "Subtotal")))
        orders.append({
            "order_number": number,
            "order_placed_date": parse_date(cell(row, "Order Date")),
            "grand_total": money_str(cell(row, "Order Total")),
            "subtotal": subtotal,
            "tax": money_str(cell(row, "Tax")),
            "shipping": _sum_str(cell(row, "Delivery Charges"), cell(row, "Bag Fee")),
            "savings": money_str(cell(row, "Savings")),
            "refund_total": money_str(cell(row, "Refund")),
            # `Payment Method` is deliberately NOT read. The export spells it
            # "Visa ending in 1840" — a card last-4, which the schema promises
            # this connector does not keep, and which nothing downstream reads.
            # A field no one consumes is one that can only ever leak.
            "payment_method": None,
            # The sheet's own count, kept as published. It counts restatement
            # rows, so it can exceed the lines stored — that disagreement is a
            # fact about the export and is not worth papering over.
            "item_count": cell(row, "Items"),
            "channel": _CHANNELS.get(_text(cell(row, "Order Type"))),
            "cancelled": bool(found.get("all_dead")),
            # This IS the detail: the export carries priced lines, so there is
            # nothing for `backfill` to fetch later. Without this flag
            # `store.store_orders` writes the order and silently skips its items.
            "detail_fetched": True,
            "items": found.get("items", []),
            "source": "xlsx",
        })
    return orders


def load(path: str | Path) -> list[dict]:
    """A Walmart order-export workbook → entity dicts for `store.store_orders`.

    Raises `WalmartParseError` rather than returning [] for a file this parser
    does not recognise. An empty result would flow into `store.assert_not_vacuous`
    and be reported as "no orders", which is a lie about a file that plainly has
    some.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:                                  # pragma: no cover
        raise WalmartParseError(
            "reading an .xlsx export needs openpyxl — run `uv sync`") from e

    path = Path(path).expanduser()
    if not path.exists():
        raise WalmartParseError(f"no such file: {path}")
    # `data_only` reads cached values instead of formulas; hyperlinks need the
    # full (non-read-only) model, which is why this is not a streaming load.
    book = load_workbook(path, data_only=True)
    for sheet in (ORDER_SHEET, ITEM_SHEET):
        if sheet not in book.sheetnames:
            raise WalmartParseError(
                f"workbook has no '{sheet}' sheet (found {book.sheetnames}) — "
                f"this does not look like a Walmart order export")

    orders = _read_orders(book[ORDER_SHEET], _read_items(book[ITEM_SHEET]))
    if not orders:
        raise WalmartParseError(
            f"'{ORDER_SHEET}' sheet has a valid header but no order rows")
    return orders


def summarize(orders: list[dict]) -> dict:
    """What an import would write, and how well it ties out.

    `reconciling` counts orders whose lines sum to the subtotal within a cent.
    It is reported rather than enforced — see the module docstring — and it is
    the number to watch across future exports: a sharp drop means the format
    moved and the item-selection rule above stopped fitting it.
    """
    from .store import to_cents

    dates = [o["order_placed_date"] for o in orders if o["order_placed_date"]]
    reconciling = comparable = 0
    for o in orders:
        subtotal = to_cents(o.get("subtotal"), strict=False)
        if not subtotal:
            continue
        comparable += 1
        lines = sum(to_cents(i.get("line_price"), strict=False) or 0
                    for i in o["items"])
        reconciling += abs(lines - subtotal) <= 1

    channels: dict[str, int] = {}
    for o in orders:
        channels[o["channel"] or "unknown"] = channels.get(o["channel"] or "unknown", 0) + 1

    return {
        "orders": len(orders),
        "items": sum(len(o["items"]) for o in orders),
        "since": min(dates) if dates else None,
        "until": max(dates) if dates else None,
        "channels": channels,
        "reconciling": reconciling,
        "comparable": comparable,
    }


def ledger_has_walmart_charges(conn, since: str, until: str | None = None) -> bool:
    """Does the BANK think there were Walmart charges in this window?

    The ground truth the anti-vacuity check leans on. If the ledger says yes and
    the import produced nothing, the parser is broken — a fact only knowable by
    comparing against data we already trust.
    """
    like = " OR ".join("merchant_norm LIKE ?" for _ in match.MERCHANT_LIKE)
    upper = " AND posted_date <= ?" if until else ""
    row = conn.execute(
        f"""SELECT COUNT(*) AS n FROM transactions
             WHERE status='posted' AND amount_cents < 0
               AND posted_date >= ?{upper} AND ({like})""",
        (since, *((until,) if until else ()), *match.MERCHANT_LIKE)).fetchone()
    return bool(row["n"])


def store_and_match(orders: list, *, scope: str, since: str) -> dict:
    """Gate, store, match, record the run — everything after parsing.

    Written through `db.connect()` (the deterministic core's read/write handle),
    never `agent_connect()`. Walmart rows are imported facts on the same footing
    as bank transactions: the agent reads them and can never write them, because
    the authorizer denies writes to any table outside `_AGENT_WRITE_TABLES` and
    no `walmart_*` table is listed.

    `since` is the window the orders came from, and it is required. An empty
    result can only be judged against one: without it, a file that genuinely
    covers a quiet period is indistinguishable from a parser that returned
    nothing, and the gate would either abort on every legitimate empty import or
    never abort at all.

    Raises `store.SyncAborted` when the result is empty but the ledger says it
    should not be, and writes nothing in that case.
    """
    with db.connect() as conn:
        expect = ledger_has_walmart_charges(conn, since)
        # Checked BEFORE a run is opened, so a broken parse leaves no trace and
        # no half-written state — nothing began, nothing to unwind.
        store.assert_not_vacuous(conn, orders=len(orders),
                                 scope_has_known_charges=expect)
        run_id = store.start_run(conn, scope)
        n = store.store_orders(conn, orders, run_id)
        result = match.run(conn)
        store.finish_run(conn, run_id, status="success",
                         orders_seen=len(orders), orders_upserted=n["orders"],
                         items_seen=n["items"], items_upserted=n["items"])
        cov = match.coverage(conn)
        # `matched` is the TOTAL that now reconciles, not what this run added.
        # Re-importing an export you already loaded adds nothing, and printing
        # "matched 0" would read as a failed run.
        return {
            "sync_run_id": run_id, "scope": scope,
            "orders": n["orders"], "items": n["items"],
            "matched": cov["split_settlements"]["orders"],
            "new_matches": result["matched"],
            "exact": result["exact"], "split": result["split"],
            "ambiguous": len(result["ambiguous"]),
            "coverage": cov,
            "horizon": match.horizon(conn),
        }


def run_import(path: str | Path, *, on_progress=None) -> dict:
    """Read the export, store it, and reconcile — the whole command."""
    orders = load(path)
    summary = summarize(orders)
    if on_progress:
        on_progress(f"  read {summary['orders']} orders · {summary['items']} items "
                    f"({summary['since']} → {summary['until']})")
    result = store_and_match(
        orders, scope=f"xlsx-import {Path(path).name}",
        # The export's own earliest order. `store_and_match` judges an empty
        # result against this window; a file that parsed to nothing has already
        # raised above, so this only ever describes a real one.
        since=summary["since"] or "1970-01-01")
    return {**result, "summary": summary}
