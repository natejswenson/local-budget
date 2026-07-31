"""Walmart spreadsheet import — the parser and the item-selection rule.

Every workbook here is BUILT, never a copy of a real export. A real one carries
the shipping name and street address of a household, and a fixture is the one
artifact guaranteed to be read by strangers. The figures are invented too, and
chosen so the arithmetic they demonstrate still holds.

The shapes reproduced below were all read off a real export before being
reinvented at fabricated amounts: two date formats in one column, weighed goods
with fractional quantities, cancelled lines, and — the case worth the most
attention — restatement rows, which are sometimes a duplicate of a line already
present and sometimes the only listing of a real item.
"""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from local_budget import db
from local_budget.connectors.walmart import import_xlsx, store
from local_budget.connectors.walmart.parse import WalmartParseError

ORDER_COLS = ["Order Number", "Order Date", "Order Type", "Items",
              "Subtotal (Before Savings)", "Savings", "Subtotal",
              "Delivery Charges", "Bag Fee", "Tax", "Refund", "Order Total",
              "Payment Method", "Ship To"]
ITEM_COLS = ["Order Number", "Order Date", "Product Name", "Qty", "Price",
             "Status", "Order Type", "Product Link"]


def workbook(orders, items, *, order_cols=ORDER_COLS, item_cols=ITEM_COLS,
             sheets=("Orders", "Items")):
    """A two-sheet export. `items` rows may carry a trailing product URL."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheets[0]
    ws.append(order_cols)
    for row in orders:
        ws.append(row)
    wi = wb.create_sheet(sheets[1])
    wi.append(item_cols)
    for row in items:
        *values, url = row
        wi.append(values)
        if url:
            wi.cell(row=wi.max_row, column=len(item_cols)).hyperlink = url
    return wb


def written(tmp_path, wb, name="export.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return path


def order_row(number="200011111111111", date="Mar 03, 2026", kind="GLASS",
              n=1, before=10.00, savings=0, subtotal=10.00, delivery=0,
              bag=0, tax=0.70, refund="", total=10.70,
              pay="Ending in 1234", ship="A Person, 1 Example St"):
    return [number, date, kind, n, before, savings, subtotal, delivery, bag,
            tax, refund, total, pay, ship]


def item_row(number="200011111111111", date="Mar 03, 2026", title="Bread",
             qty=1, price=3.98, status="Delivered on Mar 04", kind="GLASS",
             url="https://www.walmart.com/ip/Bread-Loaf/44390948"):
    return [number, date, title, qty, price, status, kind, title, url]


# ── field parsing ────────────────────────────────────────────────────────────
@pytest.mark.parametrize("raw,iso", [
    ("Mar 03, 2026", "2026-03-03"),
    # In-store receipts carry a suffix in the same column as online orders.
    ("Apr 16, 2026 purchase", "2026-04-16"),
    ("Dec 24, 2025 purchase", "2025-12-24"),
    ("", None), (None, None), ("not a date", None),
])
def test_parse_date_handles_both_formats(raw, iso):
    assert import_xlsx.parse_date(raw) == iso


@pytest.mark.parametrize("url,pid", [
    ("https://www.walmart.com/ip/14123550", "14123550"),
    ("https://www.walmart.com/ip/Fresh-Gala-Apple-Each/44390953", "44390953"),
    ("https://www.walmart.com/ip/Thing/123?athbdg=L1600", "123"),
    (None, None), ("https://www.walmart.com/ip/no-id-here", None),
])
def test_product_id_from_link(url, pid):
    assert import_xlsx.product_id(url) == pid


def test_money_is_passed_on_as_a_two_place_decimal_string():
    """The strict converter downstream reads what a human read, not a float repr."""
    assert import_xlsx.money_str(5.6) == "5.60"
    assert import_xlsx.money_str(0) == "0.00"
    assert import_xlsx.money_str("") is None
    assert import_xlsx.money_str(None) is None


def test_quantity_keeps_weights_and_defaults_to_one():
    assert import_xlsx.quantity(0.514) == 0.514      # half a pound of deli meat
    assert import_xlsx.quantity(2) == 2
    assert import_xlsx.quantity("") == 1             # the export omitted it
    assert import_xlsx.quantity(None) == 1


# ── the item-selection rule ──────────────────────────────────────────────────
def _statuses(rows):
    return import_xlsx.select_items(rows)


def raw(title, price, status):
    return {"title": title, "price": price, "status": status, "qty": 1, "url": None}


def test_cancelled_and_unavailable_lines_are_dropped():
    kept = _statuses([raw("Milk", 3.98, "Delivered"),
                      raw("Eggs", 2.17, "Canceled"),
                      raw("Cheese", 1.97, "Unavailable")])
    assert [r["title"] for r in kept] == ["Milk"]


def test_a_restatement_of_a_listed_line_is_dropped():
    """`Shopped` repeating a line already present is one purchase, not two."""
    kept = _statuses([raw("Beef", 8.97, "Picked up on Apr 16"),
                      raw("Apple", 2.21, "Picked up on Apr 16"),
                      raw("Beef", 8.97, "Shopped"),
                      raw("Apple", 2.21, "3 weight adjusted")])
    assert [r["title"] for r in kept] == ["Beef", "Apple"]
    assert sum(r["price"] for r in kept) == pytest.approx(11.18)


def test_a_restatement_of_nothing_is_a_real_item_and_survives():
    """The same status also marks the ONLY listing of an item.

    Blanket-dropping `Shopped` would silently delete real purchases — this is
    why the rule is conditional on the line it restates being present.
    """
    kept = _statuses([raw("Beef", 8.97, "Picked up on Apr 16"),
                      raw("Beef", 8.97, "Shopped"),
                      raw("Tortillas", 3.48, "Shopped")])
    assert sorted(r["title"] for r in kept) == ["Beef", "Tortillas"]


def test_two_genuine_copies_are_not_collapsed_by_one_restatement():
    """Suppression is one-for-one, so buying the same thing twice still counts twice."""
    kept = _statuses([raw("Corn", 1.56, "Picked up on Apr 16"),
                      raw("Corn", 1.56, "Picked up on Apr 16"),
                      raw("Corn", 1.56, "Shopped")])
    assert len(kept) == 2


# ── whole-file load ──────────────────────────────────────────────────────────
def test_load_maps_orders_and_items(tmp_path):
    wb = workbook(
        [order_row(n=2, before=12.95, savings=-1.00, subtotal=11.95,
                   delivery=2.99, bag=0.10, tax=0.80, total=15.74)],
        [item_row(title="Bread", price=3.98),
         item_row(title="Deli Turkey", qty=0.514, price=8.97,
                  url="https://www.walmart.com/ip/Turkey/555")])
    orders = import_xlsx.load(written(tmp_path, wb))

    assert len(orders) == 1
    o = orders[0]
    assert o["order_placed_date"] == "2026-03-03"
    assert o["channel"] == "online"
    assert o["grand_total"] == "15.74"
    assert o["subtotal"] == "12.95"       # prefers the pre-savings figure
    assert o["shipping"] == "3.09"        # delivery + bag fee, summed here
    assert o["source"] == "xlsx"
    assert o["detail_fetched"] is True    # or store.py would skip the items
    assert [i["title"] for i in o["items"]] == ["Bread", "Deli Turkey"]
    assert o["items"][1]["quantity"] == 0.514
    assert o["items"][1]["product_id"] == "555"


def test_in_store_orders_are_channelled_so_the_matcher_can_filter(tmp_path):
    wb = workbook([order_row(kind="IN_STORE", date="Apr 16, 2026 purchase")],
                  [item_row(kind="IN_STORE", date="Apr 16, 2026 purchase")])
    o = import_xlsx.load(written(tmp_path, wb))[0]
    assert o["channel"] == "in-store"
    assert o["order_placed_date"] == "2026-04-16"


def test_the_shipping_address_is_never_read(tmp_path):
    """`Ship To` holds a real name and street address. Nothing may carry it out."""
    wb = workbook([order_row(ship="Jane Doe, 100 Private Rd, Somewhere, MN 56000")],
                  [item_row()])
    o = import_xlsx.load(written(tmp_path, wb))[0]
    assert "Private Rd" not in repr(o)
    assert "Jane Doe" not in repr(o)
    # `shipping` is money — the delivery charge. What must not exist is anywhere
    # to put a recipient.
    assert not any(k in o for k in ("ship_to", "address", "recipient"))


def test_an_order_whose_every_line_was_cancelled_is_marked_cancelled(tmp_path):
    wb = workbook([order_row()], [item_row(status="Canceled")])
    o = import_xlsx.load(written(tmp_path, wb))[0]
    assert o["cancelled"] is True
    assert o["items"] == []


@pytest.mark.parametrize("kwargs,message", [
    ({"sheets": ("Purchases", "Items")}, "no 'Orders' sheet"),
    ({"order_cols": [c for c in ORDER_COLS if c != "Order Total"]},
     "missing column"),
    ({"item_cols": [c for c in ITEM_COLS if c != "Price"]}, "missing column"),
])
def test_a_changed_format_fails_loudly_rather_than_yielding_nothing(
        tmp_path, kwargs, message):
    """A silent zero is the failure this connector exists to refuse."""
    wb = workbook([order_row()], [item_row()], **kwargs)
    with pytest.raises(WalmartParseError, match=message):
        import_xlsx.load(written(tmp_path, wb))


def test_a_header_with_no_rows_is_an_error_not_an_empty_import(tmp_path):
    with pytest.raises(WalmartParseError, match="no order rows"):
        import_xlsx.load(written(tmp_path, workbook([], [])))


def test_missing_file(tmp_path):
    with pytest.raises(WalmartParseError, match="no such file"):
        import_xlsx.load(tmp_path / "absent.xlsx")


# ── summary ──────────────────────────────────────────────────────────────────
def test_summarize_counts_orders_whose_lines_tie_out(tmp_path):
    wb = workbook(
        [order_row(number="1", before=6.15, subtotal=6.15, total=6.15),
         order_row(number="2", before=10.00, subtotal=10.00, total=10.00)],
        [item_row(number="1", title="Bread", price=3.98),
         item_row(number="1", title="Milk", price=2.17),
         # Order 2's lines fall short of its subtotal, as the source's often do.
         item_row(number="2", title="Cheese", price=1.97)])
    s = import_xlsx.summarize(import_xlsx.load(written(tmp_path, wb)))
    assert (s["orders"], s["items"]) == (2, 3)
    assert (s["reconciling"], s["comparable"]) == (1, 2)
    assert s["channels"] == {"online": 2}


# ── round trip into the database ─────────────────────────────────────────────
@pytest.fixture()
def conn(tmp_path, monkeypatch):
    monkeypatch.setenv("LOCAL_BUDGET_DATA_DIR", str(tmp_path))
    dbp = tmp_path / "budget.db"
    db.init_schema(dbp)
    with db.connect(dbp) as c:
        yield c


def test_a_fractional_quantity_survives_the_round_trip(conn, tmp_path):
    """Regression: this used to store as 0 — the line kept its price and lost
    its quantity, so a report could show money spent on nothing."""
    wb = workbook([order_row(total=8.97)],
                  [item_row(title="Deli Turkey", qty=0.514, price=8.97)])
    orders = import_xlsx.load(written(tmp_path, wb))
    run_id = store.start_run(conn, "test")
    store.store_orders(conn, orders, run_id)

    row = conn.execute("SELECT quantity, line_price_cents FROM walmart_items").fetchone()
    assert row["quantity"] == pytest.approx(0.514)
    assert row["line_price_cents"] == 897


def test_the_item_sum_and_source_are_recorded(conn, tmp_path):
    """The lines do not always agree with the subtotal, so the gap is stored
    rather than left for a reader to assume away."""
    wb = workbook([order_row(before=10.00, subtotal=10.00, total=10.70)],
                  [item_row(title="Bread", price=3.98),
                   item_row(title="Milk", price=2.17)])
    orders = import_xlsx.load(written(tmp_path, wb))
    store.store_orders(conn, orders, store.start_run(conn, "test"))

    row = conn.execute("SELECT item_sum_cents, subtotal_cents, source "
                       "FROM walmart_orders").fetchone()
    assert row["item_sum_cents"] == 615
    assert row["subtotal_cents"] == 1000
    assert row["source"] == "xlsx"


def test_import_upserts_over_a_scraped_order_without_orphaning_it(conn, tmp_path):
    """The scraper and the export both write orders. Re-importing must update
    the row in place — an order number is the key matches hang off."""
    scraped = {"order_number": "200011111111111", "order_placed_date": "2026-03-03",
               "grand_total": "10.70", "channel": "online", "detail_fetched": False,
               "items": []}
    store.store_orders(conn, [scraped], store.start_run(conn, "scrape"))
    assert conn.execute("SELECT source FROM walmart_orders").fetchone()["source"] == "scrape"

    wb = workbook([order_row(total=10.70)], [item_row(title="Bread", price=3.98)])
    store.store_orders(conn, import_xlsx.load(written(tmp_path, wb)),
                       store.start_run(conn, "xlsx"))

    rows = conn.execute("SELECT source, detail_fetched FROM walmart_orders").fetchall()
    assert len(rows) == 1
    assert rows[0]["source"] == "xlsx"
    assert rows[0]["detail_fetched"] == 1
    assert conn.execute("SELECT COUNT(*) n FROM walmart_items").fetchone()["n"] == 1
